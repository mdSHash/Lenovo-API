import requests
import json
from datetime import datetime
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Function to get the machine type from the serial number
def get_type_info(serial_number):
    url = f"https://pcsupport.lenovo.com/us/en/api/v4/mse/getproducts?productId={serial_number}"
    response = requests.get(url, headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"})

    try:
        response_data = response.json()
    except json.JSONDecodeError:
        print("Failed to decode JSON from response.")
        return None

    if isinstance(response_data, list) and len(response_data) > 0:
        type_info = response_data[0].get("Name", "")
    elif isinstance(response_data, dict):
        type_info = response_data.get("Name", "")
    else:
        print("Unexpected response format:", response_data)
        return None

    type_number = None
    type_name = None
    type_matches = re.findall(r'Type (\w+)', type_info)
    if type_matches:
        type_number = type_matches[-1]

    type_name = type_info.split(' - ')[0] if ' - ' in type_info else None

    return {
        "SerialNumber": serial_number,
        "FullType": type_info,
        "TypeNumber": type_number,
        "TypeName": type_name
    }

# Function to get warranty details using the serial number and machine type
def get_warranty_info(serial_number, machine_type):
    url = "https://pcsupport.lenovo.com/us/en/api/v4/upsell/redport/getIbaseInfo"
    payload = json.dumps({
        "serialNumber": serial_number,
        "machineType": machine_type,
        "country": "us",
        "language": "en"
    })
    headers = {
        'accept': 'application/json, text/plain, */*',
        'content-type': 'application/json',
        'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
    }
    response = requests.post(url, headers=headers, data=payload)
    return response.json()

# Helper function to handle unsupported values
def format_for_excel(value):
    if isinstance(value, list):
        if all(isinstance(item, dict) for item in value):
            return ', '.join([str(item) for item in value])
        else:
            return ', '.join([str(item) for item in value])
    elif isinstance(value, dict):
        return json.dumps(value)
    return value if value is not None else "N/A"

# Function to write warranty details to Excel
def write_warranty_to_excel(sheet, warranty, status):
    sheet.append([
        format_for_excel(warranty.get('name', 'N/A')),
        format_for_excel(warranty.get('type', 'N/A')),
        format_for_excel(warranty.get('description', 'N/A')),
        format_for_excel(warranty.get('duration', 'N/A')),
        format_for_excel(warranty.get('startDate', 'N/A')),
        format_for_excel(warranty.get('endDate', 'N/A')),
        format_for_excel(warranty.get('deliveryTypeName', 'N/A')),
        format_for_excel(warranty.get('level', 'N/A')),
        status
    ])

    # Apply color based on status
    fill_color = "00FF00" if status == "Active" else "FF0000"  # Green for Active, Red for Expired
    for cell in sheet[sheet.max_row]:  # Apply to the last row
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

# Function to print enhanced warranty and machine information
def print_enhanced_info(warranty_details, sheet):
    machine_info = warranty_details['data']['machineInfo']
    base_warranties = warranty_details['data']['baseWarranties']
    upgrade_warranties = warranty_details['data']['upgradeWarranties']

    # Write basic product information to Excel
    sheet.append([
        machine_info['serial'],
        machine_info['model'],
        machine_info['productName'],
        machine_info['buildDate'],
        machine_info['shipToCountry'],
        machine_info['status'],
        machine_info['brand'],
        machine_info['series'],
        machine_info['productImage'],
        warranty_details['data'].get('warrantyStatus', 'N/A'),
        'Yes' if warranty_details['data'].get('oow', False) else 'No'
    ])

    # Get current date for comparison
    today = datetime.now().date()

    # Process base warranties
    if base_warranties:
        for warranty in base_warranties:
            end_date = datetime.strptime(warranty.get('endDate', '9999-12-31'), '%Y-%m-%d').date()
            status = 'Active' if end_date >= today else 'Expired'
            write_warranty_to_excel(sheet, warranty, status)

    # Process upgrade warranties
    if upgrade_warranties:
        for warranty in upgrade_warranties:
            end_date = datetime.strptime(warranty.get('endDate', '9999-12-31'), '%Y-%m-%d').date()
            status = 'Active' if end_date >= today else 'Expired'
            write_warranty_to_excel(sheet, warranty, status)

    # Current warranty
    if 'currentWarranty' in warranty_details['data']:
        current_warranty = warranty_details['data']['currentWarranty']
        end_date = datetime.strptime(current_warranty.get('endDate', '9999-12-31'), '%Y-%m-%d').date()
        status = 'Active' if end_date >= today else 'Expired'
        write_warranty_to_excel(sheet, current_warranty, status)

# Function to fetch part data from API
def fetch_parts_data(serial_number, machine_type):
    model_url = 'https://pcsupport.lenovo.com/us/en/api/v4/upsellAggregation/parts/model'
    model_payload = {
        "serialId": serial_number,
        "mtId": machine_type,
        "couponNumber": None,
        "source": None,
        "channel": "ESUPPORT_PARTS",
        "partFot": False,
        "firstGenCPU": False,
        "fetchProcessor": False
    }

    # Fetch data from the model API
    model_data = fetch_data(model_url, model_payload)

    asbuilt_url = 'https://pcsupport.lenovo.com/us/en/api/v4/upsellAggregation/parts/asBuilt'
    asbuilt_payload = model_payload

    # Fetch data from the asBuilt API
    asbuilt_data = fetch_data(asbuilt_url, asbuilt_payload)

    # Ensure both model_data and asbuilt_data are lists
    model_data = model_data if isinstance(model_data, list) else []
    asbuilt_data = asbuilt_data if isinstance(asbuilt_data, list) else []

    # Combine the data from both sources
    combined_data = model_data + asbuilt_data

    return combined_data

# Function to fetch data from API
def fetch_data(url, payload):
    response = requests.post(url, json=payload)
    return response.json().get('data', [])


def get_unique_color(index):
    # Generate colors in a visually distinguishable way by cycling through a set of colors
    colors = [
        "FFDDC1", "C1FFD7", "C1C1FF", "FFC1E1", "FFD7C1",
        "C1FFFA", "D7C1FF", "C1FFDD", "FFC1C1", "C1F5FF"
    ]
    return colors[index % len(colors)]

def color_by_commodity(sheet, parts_by_commodity):
    color_mapping = {}  # Store colors for each commodity value
    color_index = 0

    for commodity, parts in parts_by_commodity.items():
        # Assign a unique color for this commodity value if it doesn't have one yet
        if commodity not in color_mapping:
            color_mapping[commodity] = get_unique_color(color_index)
            color_index += 1

        fill_color = color_mapping[commodity]
        for row in parts:
            for cell in row:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

# Example usage
serial_number = input("Enter the serial number: ")
type_info = get_type_info(serial_number)

if type_info and type_info["TypeNumber"]:
    machine_type = type_info["TypeNumber"]
    warranty_details = get_warranty_info(serial_number, machine_type)

    # Sanitize the serial number for the filename
    sanitized_serial = re.sub(r'[<>:"/\\|?*]', '_', serial_number)
    excel_filename = f'Warranty_and_Parts_info_{sanitized_serial}.xlsx'

    # Create an Excel workbook and sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Warranty and Parts Info"

    # Write header for warranty information
    sheet.append([
        "Warranty Name", "Warranty Type", "Description", "Duration",
        "Start Date", "End Date", "Delivery Type", "Level", "Status"
    ])

    # Print and save enhanced warranty info
    print_enhanced_info(warranty_details, sheet)

    # Fetch and write part information
    sheet.append([])  # Add an empty row before part information
    sheet.append(["Part ID", "Part Name", "Substitutes", "Commodity Value", "Image URLs"])  # Header for parts

    parts_data = fetch_parts_data(serial_number, machine_type)
    unique_parts = {}
    for item in parts_data:
        part_id = item.get('id')
        if part_id not in unique_parts:
            unique_parts[part_id] = {
                'id': part_id,
                'name': item.get('name', 'N/A'),
                'substitutes': format_for_excel(item.get('substitutes', 'N/A')),
                'commodityVal': format_for_excel(item.get('commodityVal', 'N/A')),
                'imageUrls': format_for_excel(item.get('imageUrls', 'N/A'))
            }

    # Group rows by commodity value
    parts_by_commodity = {}
    for part in unique_parts.values():
        row = [
            format_for_excel(part['id']),
            format_for_excel(part['name']),
            format_for_excel(part['substitutes']),
            format_for_excel(part['commodityVal']),
            format_for_excel(part['imageUrls'])
        ]
        sheet.append(row)

        # Get the row just added
        row_cells = sheet[sheet.max_row]
        commodity_val = part['commodityVal']
        if commodity_val not in parts_by_commodity:
            parts_by_commodity[commodity_val] = []
        parts_by_commodity[commodity_val].append(row_cells)

    # Apply colors based on commodity value groups
    color_by_commodity(sheet, parts_by_commodity)

    # Save workbook
    workbook.save(excel_filename)
    print(f"Data saved to {excel_filename}")
else:
    print("Invalid serial number or unable to retrieve type information.")